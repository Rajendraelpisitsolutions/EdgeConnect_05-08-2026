"""
Build the QA tracking spreadsheet from the inline test-case data below.

Output: docs/qa/2026-05-27-modbus-to-opcua-pipeline-qa-tracker.xlsx
Source plan: docs/qa/2026-05-27-modbus-to-opcua-pipeline-qa-plan.md

Two sheets:
  - Cover: summary stats (Excel formulas), instructions, exit criteria
  - Test Cases: 67 test cases with dropdowns for Result + Severity,
    conditional formatting on Result.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

TEST_CASES = [
    # (TC ID, Category, Priority, Title, Pre-conditions, Steps summary, Expected)

    # FUNCTIONAL (TC-F) — 20 cases
    ("TC-F-001", "Functional", "P1", "Add Modbus source via wizard",
     "EdgeConnect + Modbus simulator running",
     "Sources tab → Add → Modbus → fill identity + connection + 3 tags → Save as draft → Config → Validate → Apply",
     "Wizard saves; source appears Running within 5s"),

    ("TC-F-002", "Functional", "P1", "Test Connection probe succeeds (valid host)",
     "Modbus simulator running on 127.0.0.1:5020",
     "Open Modbus source wizard; fill host=127.0.0.1, port=5020, unit=1; click Test Connection",
     "Inline result panel in Connection section, severity=Success, probe ID + elapsed ms shown"),

    ("TC-F-003", "Functional", "P1", "Test Connection probe fails (invalid host)",
     "—",
     "Open Modbus source wizard; host=192.0.2.1 (unreachable); click Test Connection",
     "Inline panel, severity=Error, MODBUS.CONNECT_TIMEOUT code + remediation hint"),

    ("TC-F-004", "Functional", "P1", "Add OPC UA Server destination",
     "EdgeConnect running",
     "Destinations tab → Add → OPC UA Server → fill endpoint URL + app URI → Security=None → Save → Apply",
     "Destination Running; netstat shows port 4840 listening"),

    ("TC-F-005", "Functional", "P1", "Create route wiring source → sink",
     "TC-F-001 + TC-F-004 done",
     "Routes tab → Add → pick source + sink → Save → Apply",
     "Route appears Running, no errors in diagnostics"),

    ("TC-F-006", "Functional", "P1", "Tags visible in UaExpert",
     "Pipeline configured per F-001..F-005",
     "Open UaExpert; add server opc.tcp://localhost:4840/edgeconnect; Security=None; Anonymous; Connect; browse Objects → EdgeConnect → plc → modbus-sim-1",
     "All 3 tags appear with correct OPC UA datatypes"),

    ("TC-F-007", "Functional", "P1", "Tag values update live in UaExpert",
     "TC-F-006 done",
     "Drag tags to Data Access View; observe 60s",
     "Values update at ~scan-rate cadence; no Bad quality; timestamps present"),

    ("TC-F-008", "Functional", "P2", "All 4 register classes decode correctly",
     "—",
     "Configure source with tags from Coil, DI, HR, IR classes; wire to OPC UA; subscribe in UaExpert",
     "All 4 tags correct; bools = Boolean datatype (not Int32)"),

    ("TC-F-009", "Functional", "P2", "Byte order variations decode correctly",
     "—",
     "Configure tags with AB, ABCD, CDAB, BADC orders; subscribe in UaExpert; cross-check with pymodbus snippet",
     "parts_count = 1,234,567; floats not NaN; values match raw reads"),

    ("TC-F-010", "Functional", "P2", "Scale + offset applied to temperature",
     "Simulator running (raw IR:0 = 420)",
     "Add temperature tag with scale=0.1; subscribe in UaExpert",
     "Value shows as 42.0 (NOT 420). Change scale to 1.0 → value becomes 420.0"),

    ("TC-F-011", "Functional", "P2", "String tags decoded",
     "—",
     "Add 'mode' (string16) and 'part_name' (string8); subscribe",
     "mode='AUTO' (possibly padded); part_name='SHAFT-7X'; OPC UA datatype=String"),

    ("TC-F-012", "Functional", "P1", "Edit source via pencil icon",
     "modbus-sim-1 exists",
     "Sources tab → pencil → wizard pre-fills → change PollIntervalMs 1000→500 → Save changes",
     "Navigates to source detail; snackbar confirms; UaExpert update frequency doubles"),

    ("TC-F-013", "Functional", "P2", "Edit sink via pencil icon",
     "opcua-edge exists",
     "Destinations tab → pencil → change RootFolder → Save changes",
     "UaExpert rebrowse shows new RootFolder name; old tree gone"),

    ("TC-F-014", "Functional", "P2", "Edit route via pencil icon",
     "Route exists",
     "Routes tab → pencil → verify InstanceId disabled → Cancel",
     "RouteId disabled with helper text; Cancel discards everything"),

    ("TC-F-015", "Functional", "P1", "Disable source stops data flow",
     "Pipeline running",
     "Sources tab → Disable on modbus-sim-1 → confirm",
     "Within 2s: State=Stopped; UaExpert tag quality goes BadNoCommunication/Uncertain"),

    ("TC-F-016", "Functional", "P1", "Re-enable source resumes flow",
     "TC-F-015 done",
     "Click Enable on the row",
     "Within 5s: State=Running; UaExpert tags resume updating"),

    ("TC-F-017", "Functional", "P2", "Fanout to multiple sinks",
     "Mosquitto running on localhost:1883",
     "Add MQTT sink; edit route → add MQTT to sinks (multi-select) → Apply; subscribe via UaExpert AND mosquitto_sub",
     "Both subscribers receive independently; killing one doesn't affect the other"),

    ("TC-F-018", "Functional", "P2", "Validation banner shows errors before save",
     "—",
     "Open any wizard; clear InstanceId; observe banner above sections",
     "WizardValidationBanner appears, severity=Error, lists missing field"),

    ("TC-F-019", "Functional", "P2", "Validation banner scroll-to-field",
     "—",
     "Open wizard with errors; scroll past Section 4; click a message in banner",
     "Page smoothly scrolls to referenced field; field gains focus"),

    ("TC-F-020", "Functional", "P2", "StaleEditWarningBanner on concurrent edit",
     "Two browser tabs",
     "Both tabs at /destinations/opcua-edge/edit. Tab 1: change field + Save. Tab 2: change different field + Save",
     "Tab 1 succeeds; Tab 2 shows StaleEditWarningBanner with versions + formatted timestamp + Reload"),

    # CONFIGURATION (TC-C) — 7 cases
    ("TC-C-001", "Configuration", "P1", "Draft → Validate → Apply → Rollback cycle",
     "—",
     "Save 3 separate drafts; Validate+Apply each; click Rollback on second-most-recent",
     "Each Apply increments version id; Rollback creates draft; applying it reverts state"),

    ("TC-C-002", "Configuration", "P2", "Apply rejects invalid config",
     "—",
     "Craft draft with invalid value (e.g. BrokerPort=99999); click Validate",
     "Validation fails with clear field+constraint message; Apply stays disabled"),

    ("TC-C-003", "Configuration", "P2", "Configuration history",
     "≥5 applies done",
     "Config → History tab",
     "All versions listed with timestamps, version ids, actor names; each clickable to view contents"),

    ("TC-C-004", "Configuration", "P2", "Backup export",
     "Pipeline configured",
     "Config → Export backup → download",
     "Archive contains gateway.json + history; NO plaintext secrets/passwords"),

    ("TC-C-005", "Configuration", "P2", "Backup restore",
     "Backup from TC-C-004 exists",
     "Delete sources/sinks; restore from backup",
     "Original config reproduced bit-for-bit (modulo timestamps); data flow resumes"),

    ("TC-C-006", "Configuration", "P3", "Filter rules drop expected tags",
     "Route with filter that excludes parts_count",
     "Apply route; subscribe in UaExpert",
     "All other tags update; parts_count does not appear or never updates"),

    ("TC-C-007", "Configuration", "P3", "Transform deadband suppresses small changes",
     "Deadband threshold=5 on spindle_load",
     "Apply route; observe in UaExpert",
     "Small fluctuations (<5) suppressed; large jumps (≥5) propagate"),

    # PERFORMANCE (TC-P) — 5 cases
    ("TC-P-001", "Performance", "P1", "Baseline 10 tags @ 1s scan rate",
     "10-tag source + 1 sink + 1 UaExpert subscriber",
     "Run 1 hour; capture CPU/RAM/poll counts every 5min",
     "CPU<5%; RAM growth<50MB; drop rate=0"),

    ("TC-P-002", "Performance", "P2", "50 tags @ 500ms scan rate",
     "50-tag source",
     "Run 1 hour; same metrics as P-001",
     "CPU<20%; RAM growth<200MB; drop rate=0; all 50 tags update"),

    ("TC-P-003", "Performance", "P2", "Multiple OPC UA clients",
     "5 UaExpert instances",
     "Each subscribes to all 10 tags; 30 min",
     "All 5 receive identical streams; no Bad codes; CPU stable"),

    ("TC-P-004", "Performance", "P3", "OPC UA Read latency",
     "—",
     "UaExpert Read on single tag x50; compute average + p99",
     "Average<50ms over loopback; p99<200ms"),

    ("TC-P-005", "Performance", "P3", "Subscription notification latency",
     "Sentinel write tooling",
     "Time from simulator write → UaExpert receive; p99 calc",
     "p99<2×scan interval"),

    # RELIABILITY (TC-R) — 7 cases
    ("TC-R-001", "Reliability", "P1", "Modbus simulator killed mid-stream",
     "Pipeline running",
     "Kill Python sim process; observe 60s; restart sim",
     "Within 10s: source Faulted/Reconnecting; UaExpert qual drops; within 5s of restart: Running again"),

    ("TC-R-002", "Reliability", "P2", "OPC UA Server restart preserves NodeId stability",
     "—",
     "Note NodeId in UaExpert; disable sink → drained → re-enable; note NodeId again",
     "NodeId identical before/after; programmatic pin would survive"),

    ("TC-R-003", "Reliability", "P1", "EdgeConnect service restart preserves data",
     "—",
     "Stop EdgeConnect; restart; UaExpert auto-reconnects",
     "Config preserved; flow resumes within 30s; no buffered-data loss for SAF sinks"),

    ("TC-R-004", "Reliability", "P2", "Network partition with store-and-forward (MQTT route)",
     "MQTT sink wired alongside OPC UA",
     "Stop Mosquitto; let buffer grow 5min; restart Mosquitto",
     "OPC UA unaffected; MQTT sink Buffering; buffer drains in order on restart"),

    ("TC-R-005", "Reliability", "P2", "8-hour soak test",
     "30 tags @ 500ms",
     "Run continuously 8h; hourly metrics",
     "No crashes; RAM plateau within 2h; no monotonic growth; updates remain timely"),

    ("TC-R-006", "Reliability", "P3", "24-hour soak (extended)",
     "Same as R-005",
     "Run 24h; document any memory step-changes",
     "Same as R-005 over 24h window"),

    ("TC-R-007", "Reliability", "P3", "Concurrent edits from two browsers",
     "—",
     "Covered in TC-F-020; verified here for absence of crash/corruption",
     "No data corruption, no crash; same outcome as F-020"),

    # RECOVERY (TC-RC) — 5 cases
    ("TC-RC-001", "Recovery", "P1", "Source faulted → disable → re-enable",
     "—",
     "Configure source with wrong port; goes Faulted; Disable; Edit→fix port; Enable",
     "Transitions Stopped→Starting→Running; UaExpert updates resume"),

    ("TC-RC-002", "Recovery", "P2", "Drain state during disable",
     "Slow sink or saturated route",
     "Disable mid-publishing",
     "Sink shows Draining; Action button disabled with tooltip; eventually Stopped"),

    ("TC-RC-003", "Recovery", "P2", "Apply with invalid config preserves previous",
     "—",
     "Craft broken draft; attempt Apply",
     "Apply fails with diagnostic; running config unchanged; pipeline keeps flowing"),

    ("TC-RC-004", "Recovery", "P3", "Service kill during apply",
     "—",
     "Force-kill EdgeConnect during apply window",
     "Restart picks previous OR new (never mixed); history shows clear outcome"),

    ("TC-RC-005", "Recovery", "P3", "Disk full during draft save",
     "Buffer volume near 100%",
     "Attempt draft save",
     "Save fails with clear error; existing buffered data preserved"),

    # SECURITY (TC-S) — 6 cases
    ("TC-S-001", "Security", "P1", "OPC UA Security None + Anonymous",
     "Covered by TC-F-006",
     "Re-execute F-006 with security None + anonymous",
     "Connects and serves tags"),

    ("TC-S-002", "Security", "P2", "OPC UA Sign mode shows runtime warning",
     "—",
     "OpcUa wizard → SecurityMode=Sign → Save → Apply",
     "Wizard shows release-prerequisite Warning; on apply OPCUA.SECURITY_NOT_YET_IMPLEMENTED logged"),

    ("TC-S-003", "Security", "P3", "UaExpert rejects Sign mode (when Milestone K lands)",
     "Milestone K — deferred",
     "Deferred",
     "Deferred — execute when Milestone K runtime enforcement ships"),

    ("TC-S-004", "Security", "P2", "Studio basic auth",
     "Auth.Mode=Basic + creds configured",
     "Restart EdgeConnect; open Studio",
     "Browser prompts; wrong creds=401; right creds=Studio loads; APIs require auth"),

    ("TC-S-005", "Security", "P3", "License module gating",
     "License missing modbus-tcp module",
     "Apply license; restart; attempt to add Modbus source",
     "Picker hides protocol OR save fails clearly; existing sources keep running; config changes blocked"),

    ("TC-S-006", "Security", "P2", "Secrets not exposed in backup",
     "MQTT sink with username/password",
     "Export backup; inspect archive",
     "password field redacted or omitted; no plaintext creds anywhere"),

    # DIAGNOSTICS (TC-D) — 5 cases
    ("TC-D-001", "Diagnostics", "P1", "Source state transitions visible",
     "—",
     "Stop and start simulator; watch Sources tab Status column",
     "Observed: Running→Faulted/Reconnecting→Running; transitions in detail event log"),

    ("TC-D-002", "Diagnostics", "P2", "Diagnostic events logged",
     "—",
     "Diagnostics tab → filter by source",
     "CONNECT, DISCONNECT, FAULT, RECOVER events with timestamp + code + severity"),

    ("TC-D-003", "Diagnostics", "P2", "Prometheus metrics endpoint",
     "—",
     "curl http://127.0.0.1:5080/metrics (or actual port)",
     "Standard Prometheus format; source/sink/buffer/error counters present"),

    ("TC-D-004", "Diagnostics", "P3", "Commissioning checklist",
     "—",
     "Navigate to Commissioning tab",
     "Each check ✓/✗; identity, license, ≥1 source/sink/route, etc."),

    ("TC-D-005", "Diagnostics", "P3", "Route ingest/egress counters",
     "—",
     "Routes tab → click into route detail",
     "Counters increment as data flows; reset on apply/restart per design"),

    # UX / WIZARD (TC-U) — 6 cases
    ("TC-U-001", "UX/Wizard", "P2", "All 5 protocol wizards render with WizardShell",
     "—",
     "Open each: focas2, brother-http, modbus, mqtt, opcua",
     "Header band + numbered sections + footer Cancel/Save (+ Test Connection where applicable)"),

    ("TC-U-002", "UX/Wizard", "P3", "Save button copy",
     "—",
     "Open each wizard Add and Edit",
     "Add='Save as draft'; Edit='Save changes'"),

    ("TC-U-003", "UX/Wizard", "P2", "Edit mode disables InstanceId",
     "Existing entities",
     "Enter edit mode via pencil",
     "InstanceId disabled with immutability helper text"),

    ("TC-U-004", "UX/Wizard", "P3", "Test Connection button hidden for OpcUa Server",
     "—",
     "Open OPC UA Server wizard",
     "Footer = Cancel + Save only; Info alert explains no-probe carve-out"),

    ("TC-U-005", "UX/Wizard", "P3", "Cancel discards changes",
     "—",
     "Open Add wizard; fill fields; click Cancel",
     "Navigates back; no partial config persisted"),

    ("TC-U-006", "UX/Wizard", "P3", "Browser tab close discards changes",
     "—",
     "Open Add wizard; fill; close tab; reopen",
     "Fields are default (no localStorage — ADR-0015 Rule 8)"),

    # EDGE CASES (TC-E) — 6 cases
    ("TC-E-001", "Edge", "P2", "Empty configuration / first-run state",
     "Empty gateway.json",
     "Delete config; restart; attempt to add first source",
     "Studio loads; empty state surfaces; wizard works"),

    ("TC-E-002", "Edge", "P3", "Maximum tag count",
     "—",
     "Configure source with 200+ tags; Apply; subscribe",
     "All visible; Studio responsive; perf metrics within reasonable bounds"),

    ("TC-E-003", "Edge", "P3", "Boundary values for port",
     "—",
     "Try port=1, 65535 (valid); port=0, 65536 (invalid)",
     "1 and 65535 accepted; 0 and 65536 rejected with clear error"),

    ("TC-E-004", "Edge", "P3", "Long instance ids",
     "—",
     "Try 256-char id; try id with non-regex chars",
     "Length/regex constraints enforced; clear error messages"),

    ("TC-E-005", "Edge", "P3", "Tag name with special characters",
     "—",
     "Try valid (slashes, dots); try invalid (spaces, ?)",
     "Valid charset documented + enforced; invalid rejected"),

    ("TC-E-006", "Edge", "P3", "Reload Studio while config apply in progress",
     "—",
     "Initiate apply; F5 within 1-2s",
     "Studio reloads to consistent state (pre or post apply, never mid)"),
]


# ─────────────────────────────────────────────────────────────────────────────
# Build workbook
# ─────────────────────────────────────────────────────────────────────────────

wb = Workbook()

# ─── Styles ──────────────────────────────────────────────────────────────────
FONT_NAME = "Calibri"
header_font = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", start_color="1F4E78")
title_font = Font(name=FONT_NAME, bold=True, size=16, color="1F4E78")
section_font = Font(name=FONT_NAME, bold=True, size=12, color="1F4E78")
body_font = Font(name=FONT_NAME, size=10)
small_font = Font(name=FONT_NAME, size=10, italic=True, color="595959")

thin = Side(border_style="thin", color="BFBFBF")
cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

priority_fills = {
    "P1": PatternFill("solid", start_color="FFE699"),  # warm-ish for P1
    "P2": PatternFill("solid", start_color="E2EFDA"),  # mild green for P2
    "P3": PatternFill("solid", start_color="F2F2F2"),  # neutral for P3
}

# Result conditional formatting fills
pass_fill = PatternFill("solid", start_color="C6EFCE")
fail_fill = PatternFill("solid", start_color="FFC7CE")
blocked_fill = PatternFill("solid", start_color="FFEB9C")
notrun_fill = PatternFill("solid", start_color="F2F2F2")

# ─── Sheet: Test Cases ───────────────────────────────────────────────────────
tc_sheet = wb.active
tc_sheet.title = "Test Cases"

headers = [
    "TC ID", "Category", "Priority", "Title",
    "Pre-conditions", "Steps (summary)", "Expected Result",
    "Result", "Severity", "Defect ID", "Tester", "Date Executed", "Notes",
]
for col_idx, h in enumerate(headers, start=1):
    c = tc_sheet.cell(row=1, column=col_idx, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.border = cell_border

# Write test case rows
for row_idx, tc in enumerate(TEST_CASES, start=2):
    tc_id, cat, pri, title, precond, steps, expected = tc
    row_vals = [tc_id, cat, pri, title, precond, steps, expected,
                "Not Run", "", "", "", "", ""]
    for col_idx, val in enumerate(row_vals, start=1):
        c = tc_sheet.cell(row=row_idx, column=col_idx, value=val)
        c.font = body_font
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c.border = cell_border
    # Priority colour-fill on Priority column
    pri_cell = tc_sheet.cell(row=row_idx, column=3)
    pri_cell.fill = priority_fills.get(pri, priority_fills["P3"])
    pri_cell.alignment = Alignment(horizontal="center", vertical="center")
    pri_cell.font = Font(name=FONT_NAME, size=10, bold=True)

# Column widths
col_widths = {
    "A": 12, "B": 14, "C": 10, "D": 45,
    "E": 35, "F": 60, "G": 45,
    "H": 13, "I": 11, "J": 14, "K": 14, "L": 16, "M": 35,
}
for col_letter, w in col_widths.items():
    tc_sheet.column_dimensions[col_letter].width = w

# Set row heights so wrapped text shows
tc_sheet.row_dimensions[1].height = 30
for row_idx in range(2, len(TEST_CASES) + 2):
    tc_sheet.row_dimensions[row_idx].height = 55

# Freeze header row
tc_sheet.freeze_panes = "E2"  # freeze top row + first 4 cols (TC ID..Title)

# AutoFilter on header row
last_row = 1 + len(TEST_CASES)
last_col_letter = get_column_letter(len(headers))
tc_sheet.auto_filter.ref = f"A1:{last_col_letter}{last_row}"

# ─── Data validation: Result column ──────────────────────────────────────────
dv_result = DataValidation(
    type="list",
    formula1='"Pass,Fail,Blocked,Not Run,N/A"',
    allow_blank=True,
)
dv_result.error = "Result must be one of: Pass, Fail, Blocked, Not Run, N/A"
dv_result.errorTitle = "Invalid result"
dv_result.prompt = "Pick from list"
tc_sheet.add_data_validation(dv_result)
dv_result.add(f"H2:H{last_row}")

# ─── Data validation: Severity column ────────────────────────────────────────
dv_sev = DataValidation(
    type="list",
    formula1='"S1,S2,S3,S4,S5"',
    allow_blank=True,
)
dv_sev.error = "Severity must be one of: S1, S2, S3, S4, S5"
dv_sev.errorTitle = "Invalid severity"
dv_sev.prompt = "Only fill when Result=Fail"
tc_sheet.add_data_validation(dv_sev)
dv_sev.add(f"I2:I{last_row}")

# ─── Conditional formatting: Result column ───────────────────────────────────
tc_sheet.conditional_formatting.add(
    f"H2:H{last_row}",
    CellIsRule(operator="equal", formula=['"Pass"'], fill=pass_fill),
)
tc_sheet.conditional_formatting.add(
    f"H2:H{last_row}",
    CellIsRule(operator="equal", formula=['"Fail"'], fill=fail_fill),
)
tc_sheet.conditional_formatting.add(
    f"H2:H{last_row}",
    CellIsRule(operator="equal", formula=['"Blocked"'], fill=blocked_fill),
)
tc_sheet.conditional_formatting.add(
    f"H2:H{last_row}",
    CellIsRule(operator="equal", formula=['"Not Run"'], fill=notrun_fill),
)

# ─────────────────────────────────────────────────────────────────────────────
# Sheet: Cover
# ─────────────────────────────────────────────────────────────────────────────
cover = wb.create_sheet("Cover", 0)  # insert as first sheet

# Title block
cover["A1"] = "QA Test Plan Tracker — Modbus → EdgeConnect → OPC UA Server pipeline"
cover["A1"].font = Font(name=FONT_NAME, bold=True, size=16, color="1F4E78")
cover.merge_cells("A1:F1")

cover["A2"] = "Source plan: docs/qa/2026-05-27-modbus-to-opcua-pipeline-qa-plan.md"
cover["A2"].font = small_font
cover.merge_cells("A2:F2")

cover["A3"] = "Branch: claude/m2d4-impl (post-M.2d.4)  ·  Date: 2026-05-27"
cover["A3"].font = small_font
cover.merge_cells("A3:F3")

# Summary stats section
cover["A5"] = "Summary statistics"
cover["A5"].font = section_font

# Build the stats table with Excel formulas referencing the Test Cases sheet
# Row 6: headers
stats_headers = ["Metric", "P1", "P2", "P3", "Total"]
for col_idx, h in enumerate(stats_headers, start=1):
    c = cover.cell(row=6, column=col_idx, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = cell_border

# Row 7: Total cases (formula counts on Test Cases sheet column C = Priority)
metrics = [
    ("Total cases",       'COUNTIF(\'Test Cases\'!C:C,"P1")',
                          'COUNTIF(\'Test Cases\'!C:C,"P2")',
                          'COUNTIF(\'Test Cases\'!C:C,"P3")'),
    ("Pass",              'COUNTIFS(\'Test Cases\'!C:C,"P1",\'Test Cases\'!H:H,"Pass")',
                          'COUNTIFS(\'Test Cases\'!C:C,"P2",\'Test Cases\'!H:H,"Pass")',
                          'COUNTIFS(\'Test Cases\'!C:C,"P3",\'Test Cases\'!H:H,"Pass")'),
    ("Fail",              'COUNTIFS(\'Test Cases\'!C:C,"P1",\'Test Cases\'!H:H,"Fail")',
                          'COUNTIFS(\'Test Cases\'!C:C,"P2",\'Test Cases\'!H:H,"Fail")',
                          'COUNTIFS(\'Test Cases\'!C:C,"P3",\'Test Cases\'!H:H,"Fail")'),
    ("Blocked",           'COUNTIFS(\'Test Cases\'!C:C,"P1",\'Test Cases\'!H:H,"Blocked")',
                          'COUNTIFS(\'Test Cases\'!C:C,"P2",\'Test Cases\'!H:H,"Blocked")',
                          'COUNTIFS(\'Test Cases\'!C:C,"P3",\'Test Cases\'!H:H,"Blocked")'),
    ("Not Run",           'COUNTIFS(\'Test Cases\'!C:C,"P1",\'Test Cases\'!H:H,"Not Run")',
                          'COUNTIFS(\'Test Cases\'!C:C,"P2",\'Test Cases\'!H:H,"Not Run")',
                          'COUNTIFS(\'Test Cases\'!C:C,"P3",\'Test Cases\'!H:H,"Not Run")'),
    ("N/A",               'COUNTIFS(\'Test Cases\'!C:C,"P1",\'Test Cases\'!H:H,"N/A")',
                          'COUNTIFS(\'Test Cases\'!C:C,"P2",\'Test Cases\'!H:H,"N/A")',
                          'COUNTIFS(\'Test Cases\'!C:C,"P3",\'Test Cases\'!H:H,"N/A")'),
]

start_row = 7
for i, (label, p1_f, p2_f, p3_f) in enumerate(metrics):
    r = start_row + i
    cover.cell(row=r, column=1, value=label).font = body_font
    cover.cell(row=r, column=2, value=f"={p1_f}").font = body_font
    cover.cell(row=r, column=3, value=f"={p2_f}").font = body_font
    cover.cell(row=r, column=4, value=f"={p3_f}").font = body_font
    cover.cell(row=r, column=5, value=f"=SUM(B{r}:D{r})").font = Font(name=FONT_NAME, bold=True, size=10)
    for col in range(1, 6):
        cell = cover.cell(row=r, column=col)
        cell.border = cell_border
        cell.alignment = Alignment(horizontal="center" if col > 1 else "left", vertical="center")

# Add Pass % row at the bottom (% of total)
pct_row = start_row + len(metrics)
cover.cell(row=pct_row, column=1, value="Pass %").font = Font(name=FONT_NAME, bold=True, size=10)
# Pass / Total * 100, guarded against zero (IFERROR)
for col_idx, col_letter in enumerate(["B", "C", "D", "E"], start=2):
    pass_row = start_row + 1   # row containing Pass counts
    total_row = start_row      # row containing totals
    formula = f"=IFERROR({col_letter}{pass_row}/{col_letter}{total_row},0)"
    c = cover.cell(row=pct_row, column=col_idx, value=formula)
    c.number_format = "0.0%"
    c.font = Font(name=FONT_NAME, bold=True, size=10)
    c.border = cell_border
    c.alignment = Alignment(horizontal="center", vertical="center")
cover.cell(row=pct_row, column=1).border = cell_border
cover.cell(row=pct_row, column=1).alignment = Alignment(horizontal="left", vertical="center")

# Category breakdown
breakdown_start = pct_row + 3
cover.cell(row=breakdown_start - 1, column=1, value="By category").font = section_font

cat_headers = ["Category", "Total", "Pass", "Fail", "Blocked", "Not Run", "Pass %"]
for col_idx, h in enumerate(cat_headers, start=1):
    c = cover.cell(row=breakdown_start, column=col_idx, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = cell_border

categories = ["Functional", "Configuration", "Performance", "Reliability",
              "Recovery", "Security", "Diagnostics", "UX/Wizard", "Edge"]
for i, cat in enumerate(categories):
    r = breakdown_start + 1 + i
    cover.cell(row=r, column=1, value=cat).font = body_font
    cover.cell(row=r, column=2, value=f'=COUNTIF(\'Test Cases\'!B:B,"{cat}")').font = body_font
    cover.cell(row=r, column=3, value=f'=COUNTIFS(\'Test Cases\'!B:B,"{cat}",\'Test Cases\'!H:H,"Pass")').font = body_font
    cover.cell(row=r, column=4, value=f'=COUNTIFS(\'Test Cases\'!B:B,"{cat}",\'Test Cases\'!H:H,"Fail")').font = body_font
    cover.cell(row=r, column=5, value=f'=COUNTIFS(\'Test Cases\'!B:B,"{cat}",\'Test Cases\'!H:H,"Blocked")').font = body_font
    cover.cell(row=r, column=6, value=f'=COUNTIFS(\'Test Cases\'!B:B,"{cat}",\'Test Cases\'!H:H,"Not Run")').font = body_font
    cover.cell(row=r, column=7, value=f"=IFERROR(C{r}/B{r},0)").font = body_font
    cover.cell(row=r, column=7).number_format = "0.0%"
    for col in range(1, 8):
        cell = cover.cell(row=r, column=col)
        cell.border = cell_border
        cell.alignment = Alignment(horizontal="center" if col > 1 else "left", vertical="center")

# Defect count summary (by severity)
defect_start = breakdown_start + len(categories) + 3
cover.cell(row=defect_start - 1, column=1, value="Defect summary (by severity)").font = section_font

defect_headers = ["S1 (Blocker)", "S2 (Critical)", "S3 (Major)", "S4 (Minor)", "S5 (Trivial)", "Total"]
for col_idx, h in enumerate(defect_headers, start=1):
    c = cover.cell(row=defect_start, column=col_idx, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = cell_border

# Severity counts from Test Cases sheet column I
sev_row = defect_start + 1
for col_idx, sev in enumerate(["S1", "S2", "S3", "S4", "S5"], start=1):
    formula = f'=COUNTIF(\'Test Cases\'!I:I,"{sev}")'
    c = cover.cell(row=sev_row, column=col_idx, value=formula)
    c.font = body_font
    c.border = cell_border
    c.alignment = Alignment(horizontal="center", vertical="center")
# Total
c = cover.cell(row=sev_row, column=6, value=f"=SUM(A{sev_row}:E{sev_row})")
c.font = Font(name=FONT_NAME, bold=True, size=10)
c.border = cell_border
c.alignment = Alignment(horizontal="center", vertical="center")

# Phases / instructions
inst_start = sev_row + 3
cover.cell(row=inst_start, column=1, value="Execution phases").font = section_font
phases = [
    ("Phase A — Smoke",
     "P1 cases only (~12 cases). Estimated 2 hours. Catches any blocker before deeper testing."),
    ("Phase B — Functional + Configuration",
     "All Functional + Configuration cases. Estimated 1 day."),
    ("Phase C — Reliability + Recovery + Performance",
     "Reliability (incl. 8-hour soak), Recovery, Performance. Estimated 1 day."),
    ("Phase D — Security + Diagnostics + UX + Edge",
     "Remaining categories. Estimated ½ day."),
]
for i, (name, desc) in enumerate(phases):
    r = inst_start + 1 + i
    cover.cell(row=r, column=1, value=name).font = Font(name=FONT_NAME, bold=True, size=10)
    cover.cell(row=r, column=2, value=desc).font = body_font
    cover.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)

instr_start = inst_start + len(phases) + 2
cover.cell(row=instr_start, column=1, value="How to use this tracker").font = section_font
instructions = [
    "1. Test Cases sheet: pick a row, set the Result column (dropdown: Pass/Fail/Blocked/Not Run/N/A).",
    "2. If Result=Fail, also fill Severity (S1..S5), Defect ID (link to bug tracker), and Notes.",
    "3. Tester column = your name. Date Executed = the date you ran it (YYYY-MM-DD).",
    "4. Pre-conditions and Steps are summaries — the full procedure lives in the QA plan markdown doc.",
    "5. Cover sheet auto-updates as you fill in the Test Cases sheet (Excel formulas reference column C/H/I).",
    "6. Cells use conditional formatting: Pass=green, Fail=red, Blocked=amber, Not Run=grey.",
    "7. Exit gate: 100% P1 pass (no S1/S2 open) + ≥90% P2 pass + all P3 executed.",
]
for i, line in enumerate(instructions):
    r = instr_start + 1 + i
    cover.cell(row=r, column=1, value=line).font = body_font
    cover.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)

# Cover sheet column widths
for col_letter, w in zip(["A", "B", "C", "D", "E", "F", "G"], [28, 14, 14, 14, 16, 14, 14]):
    cover.column_dimensions[col_letter].width = w

# Row heights for cover sections
cover.row_dimensions[1].height = 28
cover.row_dimensions[5].height = 22
cover.row_dimensions[breakdown_start - 1].height = 22

# Save
out_path = r"C:\dev\EdgeConnect\docs\qa\2026-05-27-modbus-to-opcua-pipeline-qa-tracker.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
print(f"Test cases: {len(TEST_CASES)}")
